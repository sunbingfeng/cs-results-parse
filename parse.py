import os.path as osp
import glob
import argparse
import numpy as np

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#### options
parser = argparse.ArgumentParser()
parser.add_argument('--report_name', type=str, required=True, help='The name of xlsx file to save to')
parser.add_argument('--result_dir', type=str, required=True, help='Path to result directory.')
parser.add_argument('--set_name', type=str, default='Set11', help='The name of the test datasets, eg: Set11, and we assumed that the result pngs are in folder of <result_dir>/<set_name>/')

args = parser.parse_args()

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, startcol=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      startcol : upper left cell column to dump data frame.
                 Per default (startcol=None) calculate the last column
                 in the existing DF and to write...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if startcol is None and sheet_name in writer.book.sheetnames:
            startcol = writer.book[sheet_name].max_column

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, **to_excel_kwargs)

    # save the workbook
    writer.save()

def parse_file_name(file_name):
    image_name = file_name.split('.')[0]

    idx = file_name.find('PSNR')
    if idx == -1:
        return []

    psnr_ssim_strs = file_name[idx:].split('_')
    psnr = float(psnr_ssim_strs[1])
    ssim = float(psnr_ssim_strs[3])

    return [image_name, psnr, ssim]


def scan_result_folder(path):
    results = []
    for file in glob.glob(path + "/" + args.set_name + "/" + "*.png"):
        file_name = osp.basename(file)
        name_without_ext = osp.splitext(file_name)[0]
        results.append(parse_file_name(name_without_ext))
    return results

def get_xlsx_sheet_size(filename, sheet_name):
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        row = writer.book[sheet_name].max_row
        col = writer.book[sheet_name].max_column
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    return [row, col]

def update_header(filename, sheet_name, header):
    [rows, cols] = get_xlsx_sheet_size(filename, sheet_name)

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        ws = writer.book[sheet_name]
        ws.cell(row=1, column=cols-1).value = header
        ws.merge_cells(start_row=1, start_column=cols-1, end_row=1, end_column=cols)
        writer.save()
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

def highlight_max(s):
    '''
    highlight the maximum in a Series yellow.
    '''
    is_max = s == s.max()
    return ['background-color: yellow' if v else '' for v in is_max]

def highlight_max(data, color='yellow'):
    '''
    highlight the maximum in a Series or DataFrame
    '''
    attr = 'background-color: {}'.format(color)
    if data.ndim == 1:  # Series from .apply(axis=0) or axis=1
        is_max = data == data.max()
        return [attr if v else '' for v in is_max]
    else:  # from .apply(axis=None)
        is_max = data == data.max().max()
        return pd.DataFrame(np.where(is_max, attr, ''),
                            index=data.index, columns=data.columns)
                                
if __name__ == "__main__":
    # global pandas options
    pd.options.display.float_format = '{:,.2f}'.format

    # parse results from folder
    results = np.array(scan_result_folder(args.result_dir))
    result_folder_name = args.result_dir.split('/')[-1]
    results = results[np.argsort(results[:, 0])]    

    avg = np.average(results[:, 1:3].astype(np.float), axis=0)
    results = np.append(results, np.array([['avg', avg[0], avg[1]]]), axis=0)

    filename = args.report_name
    sheet_name = 'Sheet1'

    if not osp.exists(filename):
        df = pd.DataFrame(results[:, 0], columns=['image'])
        df.to_excel(filename, startrow = 1, index = False)

    df = pd.DataFrame(results[:, 1:3], columns=['PSNR', 'SSIM'], dtype=np.float)
    df = df.applymap("{0:.2f}".format)

    append_df_to_excel(filename, df, startrow = 1, index=False, float_format="%.2f")
    update_header(filename, sheet_name, result_folder_name)
